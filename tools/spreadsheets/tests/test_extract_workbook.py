from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from tools.spreadsheets.extract_workbook import extract_sheet


class ExtractSheetTests(TestCase):
    def test_extracts_formulas_references_constants_and_tables(self):
        with TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "reference.xlsx"
            workbook = Workbook()
            inputs = workbook.active
            inputs.title = "Inputs"
            inputs["A1"] = 12
            costs = workbook.create_sheet("Cost Analysis")
            costs.append(["Name", "Amount"])
            costs.append(["Airframe", "='Inputs'!A1*2"])
            costs.append(["Total", "=SUM(CostTable[Amount])"])
            costs.add_table(Table(displayName="CostTable", ref="A1:B3"))
            workbook.defined_names.add(
                DefinedName("AircraftCount", attr_text="'Inputs'!$A$1")
            )
            workbook.save(workbook_path)
            workbook.close()

            manifest, golden = extract_sheet(workbook_path, "Cost Analysis")

        self.assertEqual(manifest["sheet"]["dimension"], "A1:B3")
        self.assertEqual(manifest["tables"][0]["name"], "CostTable")
        self.assertEqual(manifest["defined_names"][0]["name"], "AircraftCount")
        self.assertEqual(manifest["constants"][0]["cell"], "A1")
        self.assertEqual(manifest["formulas"][0]["cell"], "B2")
        self.assertEqual(
            manifest["formulas"][0]["references"]["cross_sheet"],
            ["'Inputs'!A1"],
        )
        self.assertEqual(
            manifest["formulas"][1]["references"]["structured"],
            ["CostTable[Amount]"],
        )
        self.assertEqual(
            [item["cell"] for item in golden["formula_results"]], ["B2", "B3"]
        )

    def test_rejects_an_unknown_sheet(self):
        with TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "reference.xlsx"
            workbook = Workbook()
            workbook.save(workbook_path)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "Unknown sheet"):
                extract_sheet(workbook_path, "Missing")
