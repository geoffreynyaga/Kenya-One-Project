"""Extract an auditable formula manifest and cached results from a workbook.

The source workbook is opened read-only from the caller's perspective: this tool
never saves it.  Formula expressions and cached values are loaded separately so
the generated reference files preserve both views of an Excel cell.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import warnings
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

SCHEMA_VERSION = 1


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def _formula_references(formula: str, sheet_name: str) -> dict[str, list[str]]:
    cross_sheet: set[str] = set()
    local_sheet: set[str] = set()
    structured: set[str] = set()
    external: set[str] = set()

    try:
        tokens = Tokenizer(formula).items
    except ValueError:
        tokens = []

    for token in tokens:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue

        reference = token.value
        if "!" in reference:
            referenced_sheet = reference.split("!", 1)[0].lstrip("(").strip("'")
            if referenced_sheet == sheet_name:
                local_sheet.add(reference)
            else:
                cross_sheet.add(reference)
        if "[" in reference and "]" in reference and "!" not in reference:
            structured.add(reference)
        if "!" in reference and reference.lstrip("(").startswith("["):
            external.add(reference)

    return {
        "cross_sheet": sorted(cross_sheet),
        "local_sheet": sorted(local_sheet),
        "structured": sorted(structured),
        "external": sorted(external),
    }


def _defined_names(workbook: Any) -> list[dict[str, Any]]:
    names = []
    for item in workbook.defined_names.values():
        names.append(
            {
                "name": item.name,
                "value": item.attr_text,
                "scope_sheet_index": item.localSheetId,
                "hidden": bool(item.hidden),
            }
        )
    return sorted(names, key=lambda item: (item["name"], item["value"] or ""))


def _tables(worksheet: Any) -> list[dict[str, Any]]:
    tables = []
    for table in worksheet.tables.values():
        tables.append(
            {
                "name": table.name,
                "display_name": table.displayName,
                "range": table.ref,
                "columns": [column.name for column in table.tableColumns],
            }
        )
    return sorted(tables, key=lambda item: item["name"])


def extract_sheet(workbook_path: Path, sheet_name: str) -> tuple[dict, dict]:
    """Return the formula manifest and golden cached results for one sheet."""

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Unable to read chart .*", category=UserWarning
        )
        formula_workbook = load_workbook(
            workbook_path, data_only=False, read_only=False
        )
        value_workbook = load_workbook(workbook_path, data_only=True, read_only=False)

    try:
        if sheet_name not in formula_workbook.sheetnames:
            available = ", ".join(formula_workbook.sheetnames)
            raise ValueError(
                f"Unknown sheet {sheet_name!r}. Available sheets: {available}"
            )

        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        formulas = []
        constants = []
        golden_results = []

        for row in formula_sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                common = {
                    "cell": cell.coordinate,
                    "number_format": cell.number_format,
                    "style_id": cell.style_id,
                }

                if cell.data_type == "f":
                    cached_value = _json_value(value_sheet[cell.coordinate].value)
                    formulas.append(
                        {
                            **common,
                            "formula": cell.value,
                            "cached_value": cached_value,
                            "references": _formula_references(cell.value, sheet_name),
                        }
                    )
                    golden_results.append(
                        {"cell": cell.coordinate, "value": cached_value}
                    )
                else:
                    constants.append(
                        {
                            **common,
                            "data_type": cell.data_type,
                            "value": _json_value(cell.value),
                        }
                    )

        extracted_content = {
            "sheet": sheet_name,
            "dimension": formula_sheet.calculate_dimension(),
            "tables": _tables(formula_sheet),
            "constants": constants,
            "formulas": formulas,
        }
        content_bytes = json.dumps(
            extracted_content, ensure_ascii=False, sort_keys=True
        ).encode()
        source = {
            "file": workbook_path.name,
            "sheet_content_sha256": hashlib.sha256(content_bytes).hexdigest(),
        }
        manifest = {
            "schema_version": SCHEMA_VERSION,
            "source": source,
            "sheet": {
                "name": sheet_name,
                "dimension": formula_sheet.calculate_dimension(),
                "merged_ranges": sorted(
                    str(cell_range) for cell_range in formula_sheet.merged_cells.ranges
                ),
            },
            "defined_names": _defined_names(formula_workbook),
            "tables": extracted_content["tables"],
            "constants": constants,
            "formulas": formulas,
        }
        golden = {
            "schema_version": SCHEMA_VERSION,
            "source": source,
            "sheet": sheet_name,
            "formula_results": golden_results,
        }
        return manifest, golden
    finally:
        formula_workbook.close()
        value_workbook.close()


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the source .xlsx file")
    parser.add_argument("--sheet", required=True, help="Worksheet name to extract")
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--golden", type=Path, required=True)
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    manifest, golden = extract_sheet(args.workbook, args.sheet)
    _write_json(args.manifest, manifest)
    _write_json(args.golden, golden)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
